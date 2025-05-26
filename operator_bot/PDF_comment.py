import os
import fitz  # PyMuPDF
import pandas as pd
import json
import re
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openai import AzureOpenAI

class PDFCommentExtractor:
    def __init__(self, pdf_path, config_path):
        self.pdf_path = pdf_path
        self.output_path = self._get_output_path()
        self.start_time = datetime.now()
        with open(config_path) as f:
            config = json.load(f)

        self.azure_client = AzureOpenAI(
            api_key=config["azure_api_token"],
            api_version=config["model_version"],
            azure_endpoint=config["azure_api_url"]
        )
        self.model_name = config["model_name"]

    def _get_output_path(self):
        input_dir = os.path.dirname(self.pdf_path)
        output_filename = os.path.splitext(os.path.basename(self.pdf_path))[0] + "_comments.xlsx"
        return os.path.join(input_dir, output_filename)

    def extract_comments(self):
        doc = fitz.open(self.pdf_path)
        comments_data = []

        for page_number in range(len(doc)):
            page = doc[page_number]
            for annot in page.annots() or []:
                comment = annot.info.get("content", "").strip()
                annot_type = annot.type[1] if annot.type else "Unknown"
                rect = annot.rect

                if comment:
                    surrounding_text = self._get_full_paragraph(page, rect)
                    coordinate_text = page.get_textbox(rect).strip()

                    comment_entry = {
                        "Comment": comment,
                        "Type": annot_type,
                        "Page Number": page_number + 1,
                        "Coordinates": f"({rect.x0}, {rect.y0}, {rect.x1}, {rect.y1})",
                        "Coordinate Text": coordinate_text,
                        "Surrounding Text": surrounding_text
                    }

                    if annot_type.lower() == "caret":
                        before, after = self.get_caret_surrounding_words(page, rect)
                        comment_entry["Before Word"] = before
                        comment_entry["After Word"] = after

                    comments_data.append(comment_entry)

        return comments_data

    def _get_full_paragraph(self, page, rect):
        """
        Extracts the full paragraph where the annotation (like caret) is located.
        Paragraph is approximated using the text block the annotation intersects with.
        """
        words = page.get_text("words")  # [x0, y0, x1, y1, "word", block_no, line_no, word_no]
        if not words:
            return ""

        # Find all words that intersect with the annotation rect
        matched_words = [w for w in words if fitz.Rect(w[:4]).intersects(rect)]
        if not matched_words:
            return ""

        anchor_block_no = matched_words[0][5]  # Use block number

        # Get all words from the same block (assumed paragraph)
        block_words = [w for w in words if w[5] == anchor_block_no]
        block_words.sort(key=lambda w: (w[6], w[7]))  # Sort by line and word order

        paragraph_text = " ".join(w[4] for w in block_words)
        return paragraph_text.strip()

    def get_caret_surrounding_words(self, page, rect):
        """
        Returns the word before and after the caret based on its X coordinate and nearby line.
        """
        x_caret = rect.x0
        y_top = rect.y0
        y_bottom = rect.y1

        words = page.get_text("words")
        # Filter words that are roughly on the same line (allow small vertical tolerance)
        line_words = [w for w in words if abs(w[1] - y_top) < 5 or abs(w[3] - y_bottom) < 5]

        # Sort words from left to right
        line_words.sort(key=lambda w: w[0])  # sort by x0

        before_word = ""
        after_word = ""

        for i, word in enumerate(line_words):
            word_x0, word_x1 = word[0], word[2]
            if word_x1 < x_caret:
                before_word = word[4]
            elif word_x0 > x_caret and after_word == "":
                after_word = word[4]
                break

        return before_word, after_word

    def _check_with_azure(self, comments_with_types, retries=3, delay=3):
        results = []

        for i, item in enumerate(comments_with_types, 1):
            comment = item["Comment"]
            comment_type = item["Type"]
            print(f"Checking comment {i}/{len(comments_with_types)}...")

            normalized_type = comment_type.lower().replace(" ", "")

            if normalized_type in {"text", "freetext"}:
                results.append({
                    "action": "none",
                    "action_explanation": "This is the one which needs to be checked from user end."
                })
                continue

            if normalized_type == "caret":
                results.append({
                    "action": "addition",
                    "action_explanation": "Add the given text at the marked position."
                })
                continue

            prompt = (
                f"Given the following comment and its type, identify what kind of action is needed:\n"
                f"Comment: \"{comment}\"\n"
                f"Type: \"{comment_type}\"\n\n"
                "Return only a valid JSON object with:\n"
                "- 'action': 'addition', 'replacement', or 'deletion'\n"
                "- 'action_explanation': a short explanation of what should be done (e.g., 'Add comma after word')\n"
                "If no issue exists, return: { \"action\": \"none\" }\n"
                "No extra text. Only valid JSON."
            )

            for attempt in range(retries):
                try:
                    response = self.azure_client.chat.completions.create(
                        model=self.model_name,
                        messages=[
                            {
                                "role": "system",
                                "content": (
                                    "You are a text review assistant. "
                                    "Always respond only with valid JSON in this format:\n"
                                    "{ \"action\": \"addition\" | \"replacement\" | \"deletion\" | \"none\", "
                                    "\"action_explanation\": string }\n"
                                    "Do not return any extra text or formatting."
                                )
                            },
                            {"role": "user", "content": prompt}
                        ],
                        max_tokens=200,
                        temperature=0.2
                    )

                    reply = response.choices[0].message.content.strip()

                    try:
                        result = json.loads(reply)
                    except json.JSONDecodeError:
                        match = re.search(r'\{.*\}', reply, re.DOTALL)
                        if match:
                            try:
                                result = json.loads(match.group())
                            except json.JSONDecodeError:
                                continue
                        else:
                            continue

                    if result.get("action") and result["action"] != "none":
                        if result["action"] in {"addition", "replacement", "deletion"}:
                            results.append(result)
                        else:
                            results.append({"action": "invalid", "action_explanation": "Invalid action returned"})
                    else:
                        results.append({"action": "none", "action_explanation": ""})
                    break

                except Exception as e:
                    print(f"Error (attempt {attempt+1}/{retries}): {e}")
                    if attempt < retries - 1:
                        time.sleep(delay)
                    else:
                        results.append({"action": "error", "action_explanation": str(e)})

        return results

    def save_to_excel(self, data):
        df = pd.DataFrame(data)

        wb = Workbook()
        ws = wb.active
        bold_font = Font(bold=True)

        for c_idx, column in enumerate(df.columns):
            cell = ws.cell(row=1, column=c_idx + 1, value=column)
            cell.font = bold_font

        for r_idx, row in df.iterrows():
            for c_idx, value in enumerate(row):
                ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)

        wb.save(self.output_path)
        print(f"Excel saved at: {self.output_path}")

    def run(self):
        try:
            print(f"Processing {self.pdf_path}...")
            comments = self.extract_comments()
            if comments:
                ai_results = self._check_with_azure(comments)
                for i, result in enumerate(ai_results):
                    comments[i]["Action"] = result.get("action", "")
                    comments[i]["Action Explanation"] = result.get("action_explanation", "")
                self.save_to_excel(comments)
            else:
                print("No comments found.")
        except Exception as e:
            print(f"Error during processing: {e}")


if __name__ == "__main__":
    pdf_path = "Kubasek_Essentials_2026Release_Ch01_ce.pdf"
    config_path = "config.json"

    if not os.path.isfile(pdf_path):
        print("Invalid PDF file path.")
    elif not os.path.isfile(config_path):
        print("Missing config file.")
    else:
        extractor = PDFCommentExtractor(pdf_path, config_path)
        extractor.run()
import tkinter as tk
from tkinter import filedialog, messagebox
import fitz  # PyMuPDF
import pandas as pd
import os
from datetime import datetime
import requests
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
import socket
import threading
import re

def convert_pdf_date(pdf_date):
    try:
        return datetime.strptime(pdf_date[2:16], "%Y%m%d%H%M%S")
    except ValueError:
        return None

def extract_comments_from_pdf(pdf_path):
    doc = fitz.open(pdf_path)
    comments_data = []
    images_folder = f"comment_images/{os.path.basename(pdf_path)}"
    os.makedirs(images_folder, exist_ok=True)

    for page_number in range(len(doc)):
        page = doc[page_number]
        for annot in page.annots():
            comment = annot.info.get("content", "")
            text_highlighted = page.get_text("text", clip=annot.rect)
            page_num = page_number + 1
            rect = annot.rect
            commenter_name = annot.info.get("title", "")
            # replies = annot.info.get("reply", "")
            comment_date = convert_pdf_date(annot.info.get("creationDate", ""))

            # Take screenshot of highlighted area
            pix = page.get_pixmap(clip=rect)
            img_path = f"{images_folder}/screenshot_page_{page_num}_{int(rect.x0)}_{int(rect.y0)}.png"
            pix.save(img_path)

            comments_data.append({
                "Image":"",
                "Comment": comment,
                "Text Highlighted": text_highlighted,
                "Page Number": page_num,
                "Commenter Name": commenter_name,
                # "Comment Replies": replies,
                "Filename": os.path.basename(pdf_path),
                "Comment Date": comment_date,
                "Screenshot": img_path,
            })

    print(f"Extracted {len(comments_data)} comments from the PDF.")
    return comments_data

def sanitize_string(value):
    if isinstance(value, str):
        return re.sub(r'[\x00-\x1F\x7F-\x9F]', '', value)
    return value

def save_to_excel(data, output_path):
    # Create a DataFrame from the extracted data
    df = pd.DataFrame(data)
    
    # Sanitize the DataFrame
    df = df.applymap(sanitize_string)

    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Write DataFrame to Excel with headers and make them bold
    bold_font = Font(bold=True)
    for c_idx, column in enumerate(df.columns):
        if df.columns[c_idx] == "Screenshot":
            continue
        cell = ws.cell(row=1, column=c_idx + 1, value=column)
        cell.font = bold_font
    
    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row):
            if df.columns[c_idx] == "Comment Date":
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            if df.columns[c_idx] == "Screenshot":
                continue
            ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)

    # Add images to the Excel file and adjust row height based on image height
    for r_idx, row in df.iterrows():
        image_path = row['Screenshot']
            
        # Adjust the column width based on the highest width image
        max_img_width = 0
        for r_idx, row in df.iterrows():
            image_path = row['Screenshot']
            if pd.notna(image_path):  # Check if the image path is not NaN
                img = ExcelImage(image_path)
                img.anchor = f'A{r_idx + 2}'  # Adjust the cell where the image will be placed (A1 is header)
                ws.add_image(img)
                
                # Adjust the row height based on the image height
                img_height = img.height
                ws.row_dimensions[r_idx + 2].height = img_height * 1  # Adjust the multiplier as needed
                
                # Track the maximum image width
                if img.width > max_img_width:
                    max_img_width = img.width

        # Set the column width based on the maximum image width
        ws.column_dimensions['A'].width = max_img_width * 0.14  # Adjust the multiplier as needed

    # Save the workbook
    wb.save(output_path)
    print(f"Data saved to {output_path}")

def get_username_ipaddress():
    hostname = socket.gethostname()
    ip_address = socket.gethostbyname(hostname)
    return ip_address, os.getlogin()

def process_pdf(pdf_path, project_name, status_label):
    def update_status(msg):
        status_label.config(text=msg)  # Update status in UI thread

    update_status("Processing PDF, please wait...")

    input_dir = os.path.dirname(pdf_path)
    output_filename = os.path.splitext(os.path.basename(pdf_path))[0] + "_output.xlsx"
    output_path = os.path.join(input_dir, output_filename)

    start_time = datetime.now()
    start_time_str = start_time.strftime("%Y-%m-%d %H:%M:%S")

    try:
        comments_data = extract_comments_from_pdf(pdf_path)
        error_file = False
    except Exception as e:
        update_status(f"Error extracting comments: {e}")
        print(f"Error extracting comments: {e}")
        return
    
    try:
        save_to_excel(comments_data, output_path)
        update_status(f"Comments extracted and saved to {output_path}")
    except Exception as e:
        update_status(f"Error saving to Excel: {e}")
        print(f"Error saving to Excel: {e}")
        error_file = True

    end_time = datetime.now()
    end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
    ip_address, username = get_username_ipaddress()

    tracking_url = "https://pdtdemo.straive.com/BG-Tracking/api/bg-process-tracking"
    data = json.dumps({
        "ip_address": str(ip_address),
        "user_name": str(username),
        "method": "PDF_COMMENT_EXTRACTOR_LOG_CREATION",
        "client_id": 1,
        "tracking_data": [{
            "file_name": os.path.basename(pdf_path),
            "start_time": start_time_str,
            "end_time": end_time_str,
            "project_name": project_name,
            "status": 1 if not error_file else 0,
            "unit_type": 7,
            "count": 1,
            "remarks": "Input file error" if error_file else "Success",
        }]
    })
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Basic QkctVFJBQ0tJTkctQVBJOjUxY2U5OGI3OTRhYjk5MmNlYTE4YjliY2ZjM2ViYmRk',
    }
    
    try:
        response = requests.post(tracking_url, headers=headers, data=data, timeout=10)
        if response.status_code == 200:
            update_status("Processing completed successfully.")
        else:
            update_status(f"Server error: {response.status_code}")
    except requests.exceptions.RequestException as e:
        update_status(f"Network error: {e}")
    
    images_folder = f"comment_images/{os.path.basename(pdf_path)}"
    try:
        if os.path.exists(images_folder):
            for file in os.listdir(images_folder):
                os.remove(os.path.join(images_folder, file))
            os.rmdir(images_folder)
    except Exception as e:
        update_status(f"Error cleaning up images folder: {e}")

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    entry_file_path.delete(0, tk.END)
    entry_file_path.insert(0, file_path)

def submit_form():
    project_name = entry_name.get()
    file_path = entry_file_path.get()
    
    if not project_name or not file_path:
        messagebox.showerror("Error", "Please enter project name and select a file")
        return

    threading.Thread(target=process_pdf, args=(file_path, project_name, result_label)).start()

# Create the main window
root = tk.Tk()
root.title("PDF_Comment_Extractor")

root.resizable(width=False, height=False)

# Create form elements
label_name = tk.Label(root, text="Project Name:")
entry_name = tk.Entry(root, width=30)

label_file_path = tk.Label(root, text="File Path:")
entry_file_path = tk.Entry(root, width=30)
browse_button = tk.Button(root, text="Browse", command=browse_file)

submit_button = tk.Button(root, text="Submit", command=submit_form)
result_label = tk.Label(root, text="", fg="blue")

label_name.grid(row=0, column=0, padx=10, pady=10, sticky=tk.E)
entry_name.grid(row=0, column=1, padx=10, pady=10)

label_file_path.grid(row=1, column=0, padx=10, pady=10, sticky=tk.E)
entry_file_path.grid(row=1, column=1, padx=10, pady=10)
browse_button.grid(row=1, column=2, padx=10, pady=10)

submit_button.grid(row=2, column=0, columnspan=3, pady=10)
result_label.grid(row=3, column=0, columnspan=3, pady=10)


# Center the window on the screen
root.update_idletasks()
window_width = root.winfo_width()
window_height = root.winfo_height()
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
position_top = int(screen_height / 2 - window_height / 2)
position_right = int(screen_width / 2 - window_width / 2)
root.geometry(f'{window_width}x{window_height}+{position_right}+{position_top}')

root.mainloop()
